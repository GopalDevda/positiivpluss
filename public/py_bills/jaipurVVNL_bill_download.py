from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import PyPDF2
import re
import time
import os as os
import pandas as pd
import openpyxl
import pdfquery
import glob
import os.path 
from PIL import Image
from pytesseract import pytesseract
from pathlib import Path
import shutil
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
options = webdriver.ChromeOptions()
options.add_experimental_option('prefs', {
"download.default_directory": "D:\\Dikshant_Electricity_data\\Electricity_Data\\JAIPURVVNL\\download_jaipur", #Change default directory for downloads
"download.prompt_for_download": False, #To auto download the file
"download.directory_upgrade": True,
"plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
})
driver = webdriver.Chrome(options=options)


login_path = "D:\\Dikshant_Electricity_data\\Electricity_Data\\JAIPURVVNL\\Jaipur_login.xlsx"
wb_obj = openpyxl.load_workbook(login_path)
sheet_obj = wb_obj.active
Account_no = sheet_obj.cell(row = 2, column = 1)
Password = sheet_obj.cell(row = 2, column = 2)
url='https://www.bijlimitra.com/custumerLoginPage'
driver.get(url)

# driver.find_element(by=By.XPATH, value='//*[@id="urLoginName"]').send_keys (Account_no.value)
# driver.find_element(by=By.XPATH, value='//*[@id="password"]').send_keys (Password.value)
time.sleep(5)
driver.find_element(by=By.XPATH, value='//*[@id="urLoginName"]').send_keys (Account_no.value)
driver.find_element(by=By.XPATH, value='//*[@id="password"]').send_keys (Password.value)
count = 0

while(True):
    try:
        count += 1
        driver.find_element(by=By.XPATH, value='//*[@id="refresh"]').click()
        element = driver.find_element(by=By.XPATH, value='//*[@id="mainCaptcha"]')
        element.screenshot('D:\\Dikshant_Electricity_data\\Electricity_Data\\JAIPURVVNL\\captcha\\captcha.png')
        image = Image.open("D:\\Dikshant_Electricity_data\\Electricity_Data\\JAIPURVVNL\\captcha\\captcha.png")
        path_to_tesseract = r'C:/Program Files/Tesseract-OCR/tesseract.exe'
        path_to_image = 'D:\\Dikshant_Electricity_data\\Electricity_Data\\JAIPURVVNL\\captcha\\captcha.png'
        #Point tessaract_cmd to tessaract.exe
        pytesseract.tesseract_cmd = path_to_tesseract
        #Open image with PIL
        img = Image.open(path_to_image)
        #Extract text from image
        text = pytesseract.image_to_string(img)
        print(text)
        time.sleep(2)
        driver.find_element(by=By.XPATH, value='//*[@id="txtInput"]').clear()
        driver.find_element(by=By.XPATH, value='//*[@id="txtInput"]').send_keys (text)
        driver.find_element(by=By.XPATH, value='//*[@id="checkout-form1"]/div[5]/button').click()

        break
    except :
        
        if count == 5:
            break
        continue
driver.find_element(by=By.XPATH, value='//*[@id="bot1-Msg1"]').click()
list = [
    '210422028368',
'210532033415',
'210434022284',
'210443033758',
'210462041573',
'210462034772',
'210434023875',
'210311000000',
'210411023540',
'210511044335',
'210511046215',
'210321042098',
'210321038038',
'210432025565',
'210501034405',
'210543024255',
'210517007778',
'210113042044',
'210113036405',
'210181025474',
'210162042158',
'210131050041',
'210462049660',
'210424029836'
]
# driver.find_element(by=By.XPATH, value='//*[@id="s2id_knumber1"]').click()
for i in list :
    driver.switch_to.window(driver.window_handles [0])
    driver.find_element(by=By.XPATH, value='//*[@id="s2id_knumber"]/a/span[2]').click()
    time.sleep(2)
    driver.find_element(by=By.XPATH, value='//*[@id="s2id_autogen1_search"]').send_keys (i)
    action = ActionChains(driver)
    action.send_keys(Keys.TAB).perform()
    time.sleep(5)
    # Select(driver.find_element(by=By.XPATH, value='//*[@id="s2id_knumber"]/a/span[2]/b')).select_by_visible_text('210113042044')

    driver.find_element(by=By.XPATH, value='//*[@id="customer-view"]/fieldset/div[2]/table/tbody/tr[1]/td[3]/button').click()
    driver.switch_to.window(driver.window_handles [1])
    time.sleep(5)
    driver.find_element(by=By.XPATH, value='//*[@id="proceed-button"]').click()
    time.sleep(5)
    driver.close()
    driver.switch_to.window(driver.window_handles [0])


print("The bills are downloaded for jaipur")



